import streamlit as st
import pandas as pd
import datetime
import plotly.express as px
import os
import io
# For Excel creation
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import requests
import base64

# Your modules
from input import process_xlsx
# main_process now returns (schedule_data, total_cost)
from main import main_process


def main():
    # 1) Basic Page Config
    st.set_page_config(page_title="Nurse Schedule Optimizer", layout="wide")

    # 2) Inject Custom CSS & Google Font
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600&display=swap');

    html, body {
      font-family: 'Inter', sans-serif;
      background-color: #f7f9fc;
      color: #333;
    }
    h1, h2, h3 {
      color: #2f4b6e; /* A professional accent color */
    }
    /* Customize buttons */
    div.stButton > button {
      background-color: #2f4b6e;
      color: #fff;
      border: none;
      padding: 0.6rem 1.2rem;
      border-radius: 5px;
    }
    div.stButton > button:hover {
      background-color: #3e5a80;
    }
    /* Header container style */
    .header {
      display: flex;
      align-items: center;
      background-color: #ffffff;
      border-bottom: 2px solid #e2e8f0;
      padding: 0.5rem 1rem;
      margin-bottom: 1rem;
    }
    .header img {
      height: 50px;
      margin-right: 1rem;
    }
    .header h1 {
      font-size: 1.5rem;
      margin: 0;
      padding: 0;
    }
    /* Footer style */
    .footer {
      background-color: #2f4b6e;
      padding: 1rem;
      text-align: center;
      color: #fff;
      margin-top: 2rem;
    }
    </style>
    """, unsafe_allow_html=True)

    # 3) Custom Header (Logo + Title)
    show_custom_header()

    # 4) Body / App Logic
    st.title("")  # We skip default Streamlit title, using our custom header

    # Radio Button: Choose Data Source
    data_source = st.radio(
        "Select Data Source",
        ["Manual Input", "Upload Excel"],
        index=0
    )

    # Show manual or upload sections
    if data_source == "Manual Input":
        show_manual_input()
    else:
        show_upload_section()

    # "Run Scheduling Algorithm" Button
    if st.button("Run Scheduling Algorithm"):
        with st.spinner("Running Scheduling..."):
            schedule_data = None
            total_cost = None  # We want to store total cost here

            if data_source == "Manual Input":
                manual_shifts = st.session_state.get("manual_shifts", [])
                manual_tasks = st.session_state.get("manual_tasks", [])
                if not manual_shifts and not manual_tasks:
                    st.error("No manual data provided!")
                    return

                tasks_dict = {str(day): [] for day in range(7)}
                for task in manual_tasks:
                    task_day = str(task["date"])
                    task_without_date = {k: v for k, v in task.items() if k != "date"}
                    tasks_dict[task_day].append(task_without_date)

                shifts_list = manual_shifts

                # main_process returns (schedule_data, total_cost)
                schedule_data, total_cost = main_process(tasks_dict, shifts_list)

            else:  # "Upload Excel"
                uploaded_file = st.session_state.get("uploaded_file", None)
                if not uploaded_file:
                    st.error("No file uploaded!")
                    return
                shifts_list, tasks_dict = process_xlsx(uploaded_file)
                # main_process => (schedule_data, total_cost)
                schedule_data, total_cost = main_process(tasks_dict, shifts_list)

        if schedule_data:
            st.success("Scheduling Completed!")
            # Show Gantt
            show_gantt_chart(schedule_data)

            # Display total cost
            if total_cost is not None:
                st.write(f"### Total Weekly Cost: {total_cost}")

            st.session_state["schedule_data"] = schedule_data
            if "gantt_excel_bytes" not in st.session_state:
                st.session_state["gantt_excel_bytes"] = create_excel_gantt_xlsx(schedule_data)
            if "shift_excel_bytes" not in st.session_state:
                st.session_state["shift_excel_bytes"] = create_shift_based_excel(schedule_data)

            st.markdown(
                download_button(
                    st.session_state["gantt_excel_bytes"],
                    "nurse_schedule.xlsx",
                    "Download Gantt Excel"
                ),
                unsafe_allow_html=True
            )

            st.markdown(
                download_button(
                    st.session_state["shift_excel_bytes"],
                    "individual_nurse_schedule.xlsx",
                    "Download Individual Nurse Schedule"
                ),
                unsafe_allow_html=True
            )

        else:
            st.error("No schedule data returned or scheduling failed.")

    # 5) Custom Footer
    show_custom_footer()


def download_button(file_data, file_name, button_text):

    b64_data = base64.b64encode(file_data).decode()

    href = f'''
        <a href="data:application/octet-stream;base64,{b64_data}" download="{file_name}"
           style="
               display: inline-block;
               padding: 0.6rem 1.2rem;
               font-size: 16px;
               color: black;
               background-color: white;
               text-align: center;
               text-decoration: none;
               border: 1px solid #d3d3d3;
               border-radius: 5px;
               margin: 10px 0;
           ">
           {button_text}
        </a>
    '''
    return href


def get_template_file() -> bytes:
    url = "https://raw.githubusercontent.com/ClancyX/2025-POBP-Nueses-Scheduling/main/Template%20with%20examples.xlsx"
    response = requests.get(url)
    if response.status_code == 200:
        return response.content
    else:
        st.error("Download Failed")
        return b""


def show_custom_header():

    """Render a custom header with the VUMC logo + title."""
    if os.path.exists("vumc_logo.png"):
        st.markdown(f"""
        <div class="header">
          <img src="data:image/png;base64,{get_base64_of_bin_file("vumc_logo.png")}" alt="Hospital Logo">
          <h1>Nurse Schedule Optimizer</h1>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="header">
          <h1>Nurse Schedule Optimizer</h1>
        </div>
        """, unsafe_allow_html=True)


def show_custom_footer():
    """Render a custom footer at the bottom."""
    st.markdown("""
    <div class="footer">
      <p>&copy; 2025 Nurse Schedule Optimizer by NW1. All rights reserved.</p>
    </div>
    """, unsafe_allow_html=True)


def get_base64_of_bin_file(bin_file):
    """Encode image file to base64 so we can embed in HTML."""
    import base64
    with open(bin_file, 'rb') as f:
        data = f.read()
    return base64.b64encode(data).decode('utf-8')


def show_manual_input():
    """Manual tasks & shifts. Store in session_state."""
    st.subheader("Manual Data Entry")

    # Shifts
    st.write("### Shifts")
    if "manual_shifts" not in st.session_state:
        st.session_state["manual_shifts"] = []

    with st.expander("Add a Shift"):
        col1, col2 = st.columns(2)
        with col1:
            s_start = st.text_input("Shift Start (HH:MM)", "08:00", key="shift_start")
            s_end = st.text_input("Shift End (HH:MM)", "16:00", key="shift_end")
            s_break_time = st.text_input("Break Time (comma-separated)", "12:00", key="shift_break")
        with col2:
            s_break_dur = st.number_input("Break Duration (min)", value=30, step=5, key="shift_break_dur")
            s_cost = st.number_input("Shift Cost", value=10.0, step=1.0, key="shift_cost")
            s_days = st.multiselect(
                "Working Days",
                options=[0, 1, 2, 3, 4, 5, 6],
                format_func=lambda x: ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"][x],
                key="s_days"
            )

        if st.button("Add Shift", key="add_shift_button"):
            new_shift = {
                "start_time": s_start,
                "end_time": s_end,
                "break_time": s_break_time,
                "break_duration": s_break_dur,
                "cost": s_cost,
                "days": ", ".join(str(d) for d in s_days)
            }
            st.session_state["manual_shifts"].append(new_shift)
            st.success("Shift Added.")

    if st.session_state["manual_shifts"]:
        df_shifts = pd.DataFrame(st.session_state["manual_shifts"])
        st.dataframe(df_shifts)
    else:
        st.info("No manual shifts yet.")

    # Tasks
    st.write("### Tasks")
    if "manual_tasks" not in st.session_state:
        st.session_state["manual_tasks"] = []

    with st.expander("Add a Task"):
        col1, col2 = st.columns(2)
        with col1:
            t_start = st.text_input("Task Start (HH:MM)", "09:00", key="task_start")
            t_end = st.text_input("Task End (HH:MM)", "10:00", key="task_end")
            t_duration = st.text_input("Task Duration (HH:MM)", "01:00", key="task_dur")
        with col2:
            t_nurses = st.number_input("Nurses Required", value=1, step=1, key="task_nurses")
            t_date = st.selectbox(
                "Select Task Day (Monday = 0, Sunday = 6)",
                options=[0, 1, 2, 3, 4, 5, 6],
                format_func=lambda x: ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"][x],
                key="t_date"
            )

        if st.button("Add Task", key="add_task_button"):
            new_task = {
                "date": t_date,
                "start_time": t_start,
                "end_time": t_end,
                "duration": t_duration,
                "nurses_required": t_nurses
            }
            st.session_state["manual_tasks"].append(new_task)
            st.success("Task Added.")

    if st.session_state["manual_tasks"]:
        df_tasks = pd.DataFrame(st.session_state["manual_tasks"])
        st.dataframe(df_tasks)
    else:
        st.info("No manual tasks yet.")


def show_upload_section():
    """File uploader for .xlsx tasks & shifts."""
    template_bytes = get_template_file()
    if template_bytes:
        # Step 2: 使用 HTML 按钮替代 `st.download_button()`
        st.markdown(
            download_button(
                template_bytes,
                "Template with examples.xlsx",
                "Download Template with examples"
            ),
            unsafe_allow_html=True
        )
    st.subheader("Upload Excel")
    file = st.file_uploader("Upload .xlsx with tasks & shifts", type=["xlsx"], key="upload_file")
    if file:
        st.session_state["uploaded_file"] = file
        st.success("File uploaded. Ready for scheduling.")
    else:
        st.session_state["uploaded_file"] = None


def show_gantt_chart(schedule_data):
    st.write("### Schedule Gantt Chart")
    gantt_data = []
    for day_key, jobs_dict in schedule_data.items():
        # day_key can be int or str
        day_index = 0
        if isinstance(day_key, int):
            day_index = day_key
        elif isinstance(day_key, str) and day_key.isdigit():
            day_index = int(day_key)

        base_day = datetime.datetime(2025, 1, 1) + datetime.timedelta(days=day_index)

        for task_id, val in jobs_dict.items():
            if len(val) < 2:
                continue
            blocks = val[0]
            if not blocks:
                continue
            start_block = min(blocks)
            end_block = max(blocks) + 1
            start_min = start_block*15
            end_min = end_block*15
            start_time = base_day + datetime.timedelta(minutes=start_min)
            end_time = base_day + datetime.timedelta(minutes=end_min)
            shift_str = str(val[1])[:30]
            gantt_data.append({
                "Task": f"Day{day_key}_Job{task_id}",
                "Start": start_time,
                "Finish": end_time,
                "Resource": shift_str
            })

    if gantt_data:
        df_gantt = pd.DataFrame(gantt_data)
        fig = px.timeline(
            df_gantt,
            x_start="Start",
            x_end="Finish",
            y="Task",
            color="Resource"
        )
        fig.update_yaxes(autorange="reversed")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("No data to display for Gantt chart.")


def convert_day_key_to_int(day_key):
    """Helper to convert day_key (int or str) to an int for sorting."""
    if isinstance(day_key, int):
        return day_key
    if isinstance(day_key, str) and day_key.isdigit():
        return int(day_key)
    return 9999


def create_excel_gantt_xlsx(schedule_data: dict) -> bytes:
    """
    Creates an Excel workbook with a matrix-like Gantt for each day.
    Rows = tasks, columns = 15-min blocks, colored cells = active intervals.
    Returns the workbook as bytes.
    """
    wb = Workbook()
    used_default_sheet = False

    day_keys_sorted = sorted(schedule_data.keys(), key=convert_day_key_to_int)
    if not day_keys_sorted:
        return b""

    for day_key in day_keys_sorted:
        day_index = convert_day_key_to_int(day_key)

        sheet_name = f"Day{day_index}"
        if not used_default_sheet:
            ws = wb.active
            ws.title = sheet_name
            used_default_sheet = True
        else:
            ws = wb.create_sheet(title=sheet_name)

        jobs_dict = schedule_data[day_key]
        all_blocks = []
        for task_data in jobs_dict.values():
            if len(task_data) >= 1 and task_data[0]:
                all_blocks.extend(task_data[0])

        if not all_blocks:
            continue

        min_block = min(all_blocks)
        max_block = max(all_blocks)

        # Row 1: time labels
        for col_i, block in enumerate(range(min_block, max_block + 1), start=2):
            minute_offset = block * 15
            hh = minute_offset // 60
            mm = minute_offset % 60
            time_label = f"{hh:02d}:{mm:02d}"
            cell = ws.cell(row=1, column=col_i)
            cell.value = time_label
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_i = 2
        sorted_job_items = sorted(jobs_dict.items(), key=lambda x: convert_day_key_to_int(x[0]))
        for task_key, val in sorted_job_items:
            ws.cell(row=row_i, column=1).value = f"Task {task_key}"
            ws.cell(row=row_i, column=1).alignment = Alignment(horizontal="left")

            if len(val) < 1:
                row_i += 1
                continue

            time_slots = val[0]
            for block in time_slots:
                col_i = (block - min_block) + 2
                fill = PatternFill("solid", fgColor=color_for_task(task_key))
                ws.cell(row=row_i, column=col_i).fill = fill

            row_i += 1

        ws.freeze_panes = "B2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def color_for_task(task_key_str: str) -> str:
    palette = [
        "FFC7CE",  # red
        "C6EFCE",  # green
        "FFEB9C",  # yellow
        "BDD7EE",  # blue
        "D6DCE5",  # gray
        "F8CBAD",  # peach
        "E2EFDA",  # mint
        "E4DFEC",  # lavender
        "FFF2CC",  # light yellow
        "DDEBF7"   # lighter blue
    ]
    try:
        if isinstance(task_key_str, int):
            t_int = task_key_str
        elif isinstance(task_key_str, str) and task_key_str.isdigit():
            t_int = int(task_key_str)
        else:
            t_int = sum(ord(c) for c in str(task_key_str))
    except:
        t_int = 999

    return palette[t_int % len(palette)]

# Individual Nurse Schedule from here


def find_available_rows(row_occupancy, start_time, end_time, required_rows):
    for row in range(2, 500):
        is_available = True
        for t in range(start_time, end_time + 1):
            if any(row + i in row_occupancy[t] for i in range(required_rows)):
                is_available = False
                break
        if is_available:
            return row
    return None


def create_shift_based_excel(schedule_data):
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    day_shift_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
    task_nurses_count = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for day_str in schedule_data:
        day = int(day_str)
        for task_id, (time_blocks, nurses_list) in schedule_data[day_str].items():
            day_blocks = [t % 96 for t in time_blocks]
            for t_day, nurses in zip(day_blocks, nurses_list):
                shift_counts = defaultdict(int)
                for shift_id in nurses:
                    shift_counts[int(shift_id)] += 1
                for shift_id, count in shift_counts.items():
                    day_shift_data[day][shift_id][t_day][task_id] = count
                    if task_nurses_count[day][shift_id][task_id] == 0:
                        task_nurses_count[day][shift_id][task_id] = count

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    for day in day_shift_data:
        for shift_id in day_shift_data[day]:
            # Get shift start and end blocks (time in 15-minute intervals)
            shift_start_block = min(day_shift_data[day][shift_id].keys())
            shift_end_block = max(day_shift_data[day][shift_id].keys())

            sheet_name = f"day_{day}_shift_{shift_id}"
            ws = wb.create_sheet(title=sheet_name)

            for col in range(shift_start_block, shift_end_block + 1):
                time_str = f"{col // 4:02d}:{(col % 4) * 15:02d}"
                cell = ws.cell(1, col - shift_start_block + 2, time_str)
                cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions[get_column_letter(col - shift_start_block + 2)].width = 7

            row_occupancy = defaultdict(set)

            for t in range(shift_start_block, shift_end_block + 1):
                tasks = day_shift_data[day][shift_id][t]
                if not tasks:
                    continue

                for task_id in tasks:
                    required_rows = task_nurses_count[day][shift_id][task_id]
                    if t == shift_start_block or task_id not in day_shift_data[day][shift_id][t - 1]:
                        merge_start = t
                        merge_end = t
                        while (merge_end + 1 <= shift_end_block and
                               task_id in day_shift_data[day][shift_id][merge_end + 1]):
                            merge_end += 1

                        start_row = find_available_rows(row_occupancy, merge_start, merge_end, required_rows)
                        if start_row is None:
                            continue

                        end_row = start_row + required_rows - 1
                        cell = ws.cell(row=start_row, column=merge_start - shift_start_block + 2)
                        cell.value = f"Task {task_id}\nNurses: {required_rows}"
                        cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                        cell.fill = PatternFill("solid", fgColor=color_for_task(str(task_id)))

                        ws.merge_cells(start_row=start_row, end_row=end_row,
                                       start_column=merge_start - shift_start_block + 2,
                                       end_column=merge_end - shift_start_block + 2)

                        for t_block in range(merge_start, merge_end + 1):
                            for row in range(start_row, end_row + 1):
                                row_occupancy[t_block].add(row)

                        for r in range(start_row, end_row + 1):
                            for c in range(merge_start, merge_end + 1):
                                ws.cell(r, c - shift_start_block + 2).border = thin_border

            max_row = max(max(rows) if rows else 1 for rows in row_occupancy.values())
            for row in range(2, max_row + 1):
                ws.cell(row, 1, f"Nurse {row - 1}")
                ws.cell(row, 1).alignment = Alignment(horizontal='right')

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


if __name__ == "__main__":
    main()
